"""
Excel Report Generator - Generowanie raportów Excel z danymi portfela
"""

from datetime import datetime
from typing import Dict, List, Any
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReportGenerator:
    """Generator raportów Excel"""
    
    def __init__(self, filename: str = None):
        """Inicjalizacja generatora"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("Pakiet openpyxl nie jest zainstalowany. Zainstaluj: pip install openpyxl")
        
        self.workbook = Workbook()
        self.filename = filename or f"raport_portfela_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Usuń domyślny sheet
        self.workbook.remove(self.workbook.active)
        
    def add_portfolio_summary(self, portfolio: Dict[str, Any]) -> None:
        """Dodaj podsumowanie portfela"""
        ws = self.workbook.create_sheet("Podsumowanie")
        
        # Nagłówek
        ws['A1'] = "RAPORT PORTFELA"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        ws['A2'] = f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=11)
        
        # Dane podsumowania
        row = 4
        self._add_summary_row(ws, row, "Wartość Portfela (PLN):", 
                            portfolio.get('PODSUMOWANIE', {}).get('Wartosc_netto_PLN', 0))
        row += 1
        
        self._add_summary_row(ws, row, "Wartość Portfela (USD):", 
                            portfolio.get('PODSUMOWANIE', {}).get('Wartosc_netto_USD', 0))
        row += 1
        
        self._add_summary_row(ws, row, "Dźwignia:", 
                            f"{portfolio.get('PODSUMOWANIE', {}).get('Leverage_ratio', 0):.1f}%")
        row += 1
        
        self._add_summary_row(ws, row, "Kurs USD/PLN:", 
                            portfolio.get('Kurs_USD_PLN', 0))
        row += 2
        
        # Sekcja akcji
        ws[f'A{row}'] = "AKCJE"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        row += 1
        
        portfel_akcji = portfolio.get('PORTFEL_AKCJI', {})
        for pozycja in portfel_akcji.get('Pozycje', [])[:10]:  # Pokaż first 10
            ticker = pozycja.get('ticker', 'N/A')
            quantity = pozycja.get('quantity', 0)
            price = pozycja.get('current_price', 0)
            value = quantity * price
            
            ws[f'A{row}'] = ticker
            ws[f'B{row}'] = quantity
            ws[f'C{row}'] = f"{price:.2f}"
            ws[f'D{row}'] = f"{value:,.2f}"
            row += 1
        
        # Sekcja krypto
        row += 1
        ws[f'A{row}'] = "KRYPTOWALUTY"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1
        
        portfel_krypto = portfolio.get('PORTFEL_KRYPTO', {})
        for crypto, data in portfel_krypto.get('pozycje', {}).items():
            ws[f'A{row}'] = crypto
            ws[f'B{row}'] = data.get('ilosc', 0)
            ws[f'C{row}'] = f"{data.get('cena_średnia', 0):.2f}"
            ws[f'D{row}'] = f"{data.get('wartosc_usd', 0):,.2f}"
            row += 1
        
        # Ustaw szerokości kolumn
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
    
    def add_stock_details(self, portfolio: Dict[str, Any]) -> None:
        """Dodaj szczegółowe dane o akcjach"""
        ws = self.workbook.create_sheet("Akcje - Szczegóły")
        
        # Nagłówki
        headers = ["Ticker", "Ilość", "Cena Aktualna", "Wartość", "% Portfela"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dane
        portfel_akcji = portfolio.get('PORTFEL_AKCJI', {})
        total_stocks_value = sum(
            p.get('quantity', 0) * p.get('current_price', 0) 
            for p in portfel_akcji.get('Pozycje', [])
        )
        
        row = 2
        for pozycja in portfel_akcji.get('Pozycje', []):
            ticker = pozycja.get('ticker', 'N/A')
            quantity = pozycja.get('quantity', 0)
            price = pozycja.get('current_price', 0)
            value = quantity * price
            percentage = (value / total_stocks_value * 100) if total_stocks_value > 0 else 0
            
            ws.cell(row=row, column=1).value = ticker
            ws.cell(row=row, column=2).value = quantity
            ws.cell(row=row, column=3).value = price
            ws.cell(row=row, column=4).value = value
            ws.cell(row=row, column=5).value = f"{percentage:.2f}%"
            
            row += 1
        
        # Ustaw szerokości kolumn
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
    
    def add_crypto_details(self, portfolio: Dict[str, Any]) -> None:
        """Dodaj szczegółowe dane o kryptowalutach"""
        ws = self.workbook.create_sheet("Krypto - Szczegóły")
        
        # Nagłówki
        headers = ["Symbol", "Ilość", "Cena Średnia", "Wartość USD", "% Portfela"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dane
        portfel_krypto = portfolio.get('PORTFEL_KRYPTO', {})
        total_crypto_value = sum(
            data.get('wartosc_usd', 0) 
            for data in portfel_krypto.get('pozycje', {}).values()
        )
        
        row = 2
        for crypto, data in portfel_krypto.get('pozycje', {}).items():
            quantity = data.get('ilosc', 0)
            avg_price = data.get('cena_średnia', 0)
            value = data.get('wartosc_usd', 0)
            percentage = (value / total_crypto_value * 100) if total_crypto_value > 0 else 0
            
            ws.cell(row=row, column=1).value = crypto
            ws.cell(row=row, column=2).value = quantity
            ws.cell(row=row, column=3).value = avg_price
            ws.cell(row=row, column=4).value = value
            ws.cell(row=row, column=5).value = f"{percentage:.2f}%"
            
            row += 1
        
        # Ustaw szerokości kolumn
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
    
    def add_performance_metrics(self, performance_data: Dict[str, Any]) -> None:
        """Dodaj metryki wydajności"""
        ws = self.workbook.create_sheet("Metryki")
        
        row = 1
        for metric, value in performance_data.items():
            ws.cell(row=row, column=1).value = metric
            ws.cell(row=row, column=2).value = value
            
            # Formatowanie
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _add_summary_row(self, ws, row: int, label: str, value: Any) -> None:
        """Pomocnik do dodawania wiersza podsumowania"""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].alignment = Alignment(horizontal="right")
    
    def save(self) -> str:
        """Zapisz raport do pliku"""
        try:
            self.workbook.save(self.filename)
            print(f"✅ Raport Excel zapisany: {self.filename}")
            return self.filename
        except Exception as e:
            print(f"❌ Błąd przy zapisywaniu raportu: {e}")
            return None


def generate_full_report(portfolio: Dict[str, Any], output_filename: str = None) -> str:
    """Wygeneruj pełny raport Excel"""
    try:
        generator = ExcelReportGenerator(output_filename)
        
        # Dodaj wszystkie arkusze
        generator.add_portfolio_summary(portfolio)
        generator.add_stock_details(portfolio)
        generator.add_crypto_details(portfolio)
        
        # Dodaj metryki wydajności
        performance_data = {
            'Łączna wartość (PLN)': portfolio.get('PODSUMOWANIE', {}).get('Wartosc_netto_PLN', 0),
            'Łączna wartość (USD)': portfolio.get('PODSUMOWANIE', {}).get('Wartosc_netto_USD', 0),
            'Dźwignia': f"{portfolio.get('PODSUMOWANIE', {}).get('Leverage_ratio', 0):.1f}%",
            'Kurs USD/PLN': portfolio.get('Kurs_USD_PLN', 0),
            'Liczba pozycji akcji': len(portfolio.get('PORTFEL_AKCJI', {}).get('Pozycje', [])),
            'Liczba pozycji krypto': len(portfolio.get('PORTFEL_KRYPTO', {}).get('pozycje', {})),
        }
        
        generator.add_performance_metrics(performance_data)
        
        # Zapisz raport
        return generator.save()
        
    except Exception as e:
        print(f"❌ Błąd przy generowaniu raportu: {e}")
        return None
